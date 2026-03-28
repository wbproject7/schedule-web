"""
직원 스케줄 웹 생성기 - 상업용 (Commercial Edition)
=====================================================
OR-Tools CP-SAT Solver 기반 스케줄 최적화 SaaS
매장별 계정 / 직원 DB / 스케줄 이력 / 인쇄 뷰

배포: Render.com (무료) + Supabase PostgreSQL (무료)
유지비용: 0원
"""

from flask import Flask, request, jsonify, send_file, send_from_directory, make_response
import pandas as pd
import os
import uuid
import time
import re
import secrets
import json
import calendar

from db import (
    init_db, IS_PG, create_store, get_store_by_code, get_store_by_id,
    update_store_settings, update_store_password, default_settings,
    get_employees, add_employee, update_employee, delete_employee,
    reactivate_employee, bulk_add_employees,
    save_schedule, get_schedules, get_schedule_by_id, delete_schedule,
    get_schedule_count, get_last_schedule, verify_password,
    save_token, get_token, delete_token, cleanup_expired_tokens,
)
from solver import solve_schedule, get_prev_month_tail
from holidays import get_holidays, get_holiday_days

app = Flask(__name__, template_folder='templates', static_folder='static')
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', secrets.token_hex(32))

# ============================================================
# 설정
# ============================================================
import tempfile as _tmpmod
OUTPUT_DIR = os.environ.get('OUTPUT_DIR', os.path.join(os.path.dirname(os.path.abspath(__file__)), 'output'))
if not os.path.exists(OUTPUT_DIR):
    OUTPUT_DIR = os.path.join(_tmpmod.gettempdir(), 'schedule_output')
os.makedirs(OUTPUT_DIR, exist_ok=True)

TOKEN_TTL = 86400 * 7  # 7일
SUPER_ADMIN_PW = os.environ.get('SUPER_ADMIN_PW', '9856')

# 파일 관리 {file_id: (path, timestamp)}
file_store = {}
FILE_TTL = 3600

# Rate limiting
rate_store = {}
RATE_WINDOW = 60
RATE_MAX = 20


# ============================================================
# 보안 유틸리티
# ============================================================

def get_current_store():
    """토큰에서 매장 ID 추출. 실패 시 None."""
    auth = request.headers.get('Authorization', '')
    if not auth.startswith('Bearer '):
        return None
    token = auth[7:]
    info = get_token(token)
    if not info or time.time() > info['expiry']:
        if info:
            delete_token(token)
        return None
    return info['store_id']


def require_auth():
    store_id = get_current_store()
    if not store_id:
        return None, (jsonify({'success': False, 'error': '인증이 필요합니다.'}), 401)
    return store_id, None


def check_rate_limit():
    ip = request.remote_addr or 'unknown'
    now = time.time()
    if ip not in rate_store:
        rate_store[ip] = []
    rate_store[ip] = [t for t in rate_store[ip] if now - t < RATE_WINDOW]
    if len(rate_store[ip]) >= RATE_MAX:
        return False
    rate_store[ip].append(now)
    return True


def cleanup_files():
    now = time.time()
    expired = [fid for fid, (path, ts) in file_store.items() if now - ts > FILE_TTL]
    for fid in expired:
        path, _ = file_store.pop(fid, (None, 0))
        if path:
            try:
                os.remove(path)
            except OSError:
                pass


def cleanup_tokens():
    cleanup_expired_tokens()


# ============================================================
# 보안 미들웨어
# ============================================================

ALLOWED_ORIGINS = os.environ.get('ALLOWED_ORIGINS', '*').split(',')

@app.before_request
def ensure_db():
    """DB 테이블이 없으면 자동 재생성"""
    if not IS_PG:
        try:
            from db import get_db
            conn = get_db()
            conn.execute('SELECT 1 FROM stores LIMIT 1')
            conn.close()
        except Exception:
            init_db()


@app.after_request
def security_headers(response):
    """보안 헤더 + CORS"""
    # CORS
    origin = request.headers.get('Origin', '')
    if '*' in ALLOWED_ORIGINS or origin in ALLOWED_ORIGINS:
        response.headers['Access-Control-Allow-Origin'] = origin or '*'
    response.headers['Access-Control-Allow-Headers'] = 'Content-Type, Authorization'
    response.headers['Access-Control-Allow-Methods'] = 'GET, POST, PUT, DELETE, OPTIONS'

    # Security headers
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['X-Frame-Options'] = 'DENY'
    response.headers['X-XSS-Protection'] = '1; mode=block'
    response.headers['Referrer-Policy'] = 'strict-origin-when-cross-origin'

    # HTTPS 전용 (프로덕션)
    if os.environ.get('RENDER') or os.environ.get('FORCE_HTTPS'):
        response.headers['Strict-Transport-Security'] = 'max-age=31536000; includeSubDomains'
        response.headers['Content-Security-Policy'] = (
            "default-src 'self'; "
            "script-src 'self' 'unsafe-inline'; "
            "style-src 'self' 'unsafe-inline'; "
            "img-src 'self' data:; "
            "connect-src 'self'; "
            "frame-ancestors 'none'"
        )

    return response


@app.route('/api/<path:path>', methods=['OPTIONS'])
def handle_options(path):
    """CORS preflight for API routes"""
    return '', 204


# ============================================================
# 라우트: 페이지
# ============================================================

@app.route('/')
def index():
    return send_from_directory('templates', 'index.html')


# ============================================================
# 라우트: 인증
# ============================================================

@app.route('/api/store/register', methods=['POST'])
def api_register():
    """매장 등록"""
    if not check_rate_limit():
        return jsonify({'success': False, 'error': '요청이 너무 많습니다.'}), 429

    data = request.json or {}
    name = (data.get('name') or '').strip()
    code = (data.get('code') or '').strip().lower()
    password = data.get('password', '')

    errors = []
    if not name or len(name) > 50:
        errors.append('매장 이름을 입력해주세요. (50자 이내)')
    if not code or len(code) < 3 or len(code) > 20 or not re.match(r'^[a-z0-9_-]+$', code):
        errors.append('매장 코드는 3~20자 영문 소문자/숫자/하이픈만 가능합니다.')
    if not password or len(password) < 4:
        errors.append('비밀번호는 4자 이상이어야 합니다.')

    if errors:
        return jsonify({'success': False, 'error': ' / '.join(errors)}), 400

    store_id = create_store(name, code, password)
    if not store_id:
        return jsonify({'success': False, 'error': '이미 사용 중인 매장 코드입니다.'}), 409

    # 자동 로그인
    token = secrets.token_hex(32)
    save_token(token, store_id, time.time() + TOKEN_TTL)

    return jsonify({
        'success': True,
        'token': token,
        'store': {'id': store_id, 'name': name, 'code': code},
    })


@app.route('/api/store/login', methods=['POST'])
def api_login():
    """매장 로그인"""
    if not check_rate_limit():
        return jsonify({'success': False, 'error': '요청이 너무 많습니다.'}), 429

    cleanup_tokens()
    data = request.json or {}
    code = (data.get('code') or '').strip().lower()
    password = data.get('password', '')

    store = get_store_by_code(code)
    if not store or not verify_password(password, store['password_hash']):
        return jsonify({'success': False, 'error': '매장 코드 또는 비밀번호가 올바르지 않습니다.'}), 401

    token = secrets.token_hex(32)
    save_token(token, store['id'], time.time() + TOKEN_TTL)

    return jsonify({
        'success': True,
        'token': token,
        'store': {
            'id': store['id'],
            'name': store['name'],
            'code': store['code'],
        },
    })


@app.route('/api/store/verify', methods=['GET'])
def api_verify():
    """토큰 유효성 확인 + 매장 정보"""
    store_id = get_current_store()
    if not store_id:
        return jsonify({'success': False}), 401
    store = get_store_by_id(store_id)
    if not store:
        return jsonify({'success': False}), 401
    settings = json.loads(store.get('settings', '{}'))
    return jsonify({
        'success': True,
        'store': {
            'id': store['id'],
            'name': store['name'],
            'code': store['code'],
            'settings': settings,
        },
    })


@app.route('/api/store/settings', methods=['PUT'])
def api_update_settings():
    """매장 설정 업데이트"""
    store_id, err = require_auth()
    if err:
        return err
    data = request.json or {}
    settings = data.get('settings', {})
    update_store_settings(store_id, settings)
    return jsonify({'success': True})


@app.route('/api/store/password', methods=['PUT'])
def api_change_password():
    """비밀번호 변경"""
    store_id, err = require_auth()
    if err:
        return err
    data = request.json or {}
    current = data.get('current', '')
    new_pw = data.get('new', '')

    store = get_store_by_id(store_id)
    if not verify_password(current, store['password_hash']):
        return jsonify({'success': False, 'error': '현재 비밀번호가 올바르지 않습니다.'}), 400
    if len(new_pw) < 4:
        return jsonify({'success': False, 'error': '새 비밀번호는 4자 이상이어야 합니다.'}), 400

    update_store_password(store_id, new_pw)
    return jsonify({'success': True})


# ============================================================
# 라우트: 직원 관리
# ============================================================

@app.route('/api/employees', methods=['GET'])
def api_get_employees():
    store_id, err = require_auth()
    if err:
        return err
    include_inactive = request.args.get('all') == '1'
    emps = get_employees(store_id, active_only=not include_inactive)
    return jsonify({'success': True, 'employees': emps})


@app.route('/api/employees', methods=['POST'])
def api_add_employee():
    store_id, err = require_auth()
    if err:
        return err
    data = request.json or {}

    # 단일 추가
    if 'name' in data:
        name = (data.get('name') or '').strip()
        if not name or len(name) > 20:
            return jsonify({'success': False, 'error': '이름을 입력해주세요. (20자 이내)'}), 400
        role = data.get('role', 'staff')
        do_count = data.get('doCount')
        emp_id = add_employee(store_id, name, role, do_count)
        if not emp_id:
            return jsonify({'success': False, 'error': f'이미 등록된 직원입니다: {name}'}), 409
        return jsonify({'success': True, 'id': emp_id})

    # 일괄 추가
    if 'names' in data:
        names = data['names']
        if isinstance(names, str):
            names = [n.strip() for n in names.split(',') if n.strip()]
        role = data.get('role', 'staff')
        added = bulk_add_employees(store_id, names, role)
        return jsonify({'success': True, 'added': added, 'count': len(added)})

    return jsonify({'success': False, 'error': '이름을 입력해주세요.'}), 400


@app.route('/api/employees/<int:emp_id>', methods=['PUT'])
def api_update_employee(emp_id):
    store_id, err = require_auth()
    if err:
        return err
    data = request.json or {}
    kwargs = {}
    if 'name' in data:
        kwargs['name'] = data['name'].strip()
    if 'role' in data:
        kwargs['role'] = data['role']
    if 'doCount' in data:
        kwargs['do_count'] = data['doCount']
    if 'sortOrder' in data:
        kwargs['sort_order'] = data['sortOrder']

    update_employee(emp_id, store_id, **kwargs)
    return jsonify({'success': True})


@app.route('/api/employees/<int:emp_id>', methods=['DELETE'])
def api_delete_employee(emp_id):
    store_id, err = require_auth()
    if err:
        return err
    delete_employee(emp_id, store_id)
    return jsonify({'success': True})


@app.route('/api/employees/<int:emp_id>/reactivate', methods=['POST'])
def api_reactivate_employee(emp_id):
    store_id, err = require_auth()
    if err:
        return err
    reactivate_employee(emp_id, store_id)
    return jsonify({'success': True})


# ============================================================
# 라우트: 공휴일
# ============================================================

@app.route('/api/holidays/<int:year>/<int:month>', methods=['GET'])
def api_holidays(year, month):
    if not (2020 <= year <= 2035 and 1 <= month <= 12):
        return jsonify({'success': False, 'error': '잘못된 연/월입니다.'}), 400
    return jsonify({
        'success': True,
        'holidays': get_holidays(year, month),
    })


# ============================================================
# 라우트: 스케줄 생성
# ============================================================

@app.route('/api/solve', methods=['POST'])
def api_solve():
    store_id, err = require_auth()
    if err:
        return err
    if not check_rate_limit():
        return jsonify({'success': False, 'error': '요청이 너무 많습니다.'}), 429

    cleanup_files()

    try:
        data = request.json
        if not data:
            return jsonify({'success': False, 'error': '요청 데이터가 비어있습니다.'}), 400

        # 입력 검증
        errors = _validate_solve_input(data)
        if errors:
            return jsonify({'success': False, 'error': ' / '.join(errors)}), 400

        year = int(data['year'])
        month = int(data['month'])
        employees = data['employees']
        holidays_list = [int(d) for d in data.get('holidays', [])]
        constraints = data['constraints']
        pre_requests = data.get('preRequests', {})
        fair_weekend = data.get('fairWeekend', True)
        managers = data.get('managers', [])
        employee_do_counts = data.get('employeeDoCountsMap', {})
        use_prev_month = data.get('usePrevMonth', False)

        # 전월 연속근무 계산
        prev_month_tail = {}
        if use_prev_month:
            prev_year = year if month > 1 else year - 1
            prev_month = month - 1 if month > 1 else 12
            prev_schedule = get_last_schedule(store_id, prev_year, prev_month)
            if prev_schedule:
                prev_month_tail = get_prev_month_tail(prev_schedule, employees)

        # 솔버 실행
        result = solve_schedule(
            employees=employees,
            year=year,
            month=month,
            holidays=holidays_list,
            constraints=constraints,
            pre_requests=pre_requests,
            fair_weekend=fair_weekend,
            managers=managers,
            employee_do_counts=employee_do_counts,
            prev_month_tail=prev_month_tail,
        )

        if not result['success']:
            return jsonify(result)

        # 파일 생성
        file_id = str(uuid.uuid4())[:8]
        excel_id, csv_id = _save_files(
            result['schedule'], employees, year, month,
            result['calendarInfo'], employee_do_counts, constraints, file_id
        )
        result['files'] = {
            'excel': f'/api/download/{excel_id}',
            'csv': f'/api/download/{csv_id}',
        }

        # DB에 저장
        sch_id = save_schedule(
            store_id=store_id,
            year=year,
            month=month,
            schedule_data=result['schedule'],
            constraints_data=constraints,
            pre_requests_data=pre_requests,
            verification_data=result['verification'],
            conflicts_data=result['conflicts'],
            file_excel=excel_id,
            file_csv=csv_id,
        )
        result['scheduleId'] = sch_id

        return jsonify(result)

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': '스케줄 생성 중 서버 오류가 발생했습니다.'}), 500


# ============================================================
# 라우트: 스케줄 이력
# ============================================================

@app.route('/api/schedules', methods=['GET'])
def api_get_schedules():
    store_id, err = require_auth()
    if err:
        return err
    limit = min(int(request.args.get('limit', 20)), 100)
    offset = int(request.args.get('offset', 0))
    schedules = get_schedules(store_id, limit, offset)
    total = get_schedule_count(store_id)
    return jsonify({'success': True, 'schedules': schedules, 'total': total})


@app.route('/api/schedules/<int:sch_id>', methods=['GET'])
def api_get_schedule(sch_id):
    store_id, err = require_auth()
    if err:
        return err
    sch = get_schedule_by_id(sch_id, store_id)
    if not sch:
        return jsonify({'success': False, 'error': '스케줄을 찾을 수 없습니다.'}), 404
    return jsonify({'success': True, 'schedule': sch})


@app.route('/api/schedules/<int:sch_id>', methods=['DELETE'])
def api_delete_schedule(sch_id):
    store_id, err = require_auth()
    if err:
        return err
    delete_schedule(sch_id, store_id)
    return jsonify({'success': True})


# ============================================================
# 라우트: 파일 다운로드
# ============================================================

@app.route('/api/download/<file_id>')
def download(file_id):
    token = request.args.get('token', '')
    info = get_token(token)
    if not info or time.time() > info['expiry']:
        if info:
            delete_token(token)
        return jsonify({'error': '인증이 필요합니다.'}), 401

    if not re.match(r'^[a-f0-9]{8}_(excel|csv)$', file_id):
        return jsonify({'error': '잘못된 파일 ID입니다.'}), 400
    if file_id not in file_store:
        return jsonify({'error': '파일을 찾을 수 없습니다.'}), 404

    path, _ = file_store[file_id]
    if not os.path.exists(path):
        return jsonify({'error': '파일이 삭제되었습니다.'}), 404
    return send_file(path, as_attachment=True)


# ============================================================
# 라우트: 대시보드 통계
# ============================================================

@app.route('/api/dashboard', methods=['GET'])
def api_dashboard():
    store_id, err = require_auth()
    if err:
        return err

    emps = get_employees(store_id)
    schedules = get_schedules(store_id, limit=5)
    total = get_schedule_count(store_id)

    return jsonify({
        'success': True,
        'employeeCount': len(emps),
        'scheduleCount': total,
        'recentSchedules': schedules,
    })


# ============================================================
# 라우트: 슈퍼관리자
# ============================================================

@app.route('/api/admin/verify', methods=['POST'])
def api_admin_verify():
    """슈퍼관리자 비밀번호 확인"""
    data = request.json or {}
    pw = data.get('password', '')
    if pw != SUPER_ADMIN_PW:
        return jsonify({'success': False, 'error': '비밀번호가 올바르지 않습니다.'}), 403
    return jsonify({'success': True})


@app.route('/api/admin/stores', methods=['GET'])
def api_admin_stores():
    """전체 매장 목록 (슈퍼관리자 전용)"""
    pw = request.headers.get('X-Admin-Password', '')
    if pw != SUPER_ADMIN_PW:
        return jsonify({'success': False, 'error': '권한이 없습니다.'}), 403

    from db import get_db, _fetchall
    conn = get_db()
    try:
        stores = _fetchall(conn,
            'SELECT id, name, code, created_at FROM stores ORDER BY id DESC')
        # 매장별 직원 수, 스케줄 수 추가
        for s in stores:
            emps = _fetchall(conn,
                'SELECT COUNT(*) as cnt FROM employees WHERE store_id = ? AND active = 1', (s['id'],))
            s['employeeCount'] = emps[0]['cnt'] if emps else 0
            schs = _fetchall(conn,
                'SELECT COUNT(*) as cnt FROM schedules WHERE store_id = ?', (s['id'],))
            s['scheduleCount'] = schs[0]['cnt'] if schs else 0
            if s.get('created_at') and not isinstance(s['created_at'], str):
                s['created_at'] = str(s['created_at'])
        return jsonify({'success': True, 'stores': stores})
    finally:
        conn.close()


@app.route('/api/admin/stores/<int:store_id>', methods=['DELETE'])
def api_admin_delete_store(store_id):
    """매장 삭제 (슈퍼관리자 전용)"""
    pw = request.headers.get('X-Admin-Password', '')
    if pw != SUPER_ADMIN_PW:
        return jsonify({'success': False, 'error': '권한이 없습니다.'}), 403

    from db import get_db, _execute
    conn = get_db()
    try:
        _execute(conn, 'DELETE FROM stores WHERE id = ?', (store_id,))
        conn.commit()
        return jsonify({'success': True})
    finally:
        conn.close()


# ============================================================
# 입력 검증
# ============================================================

def _validate_solve_input(data):
    errors = []

    try:
        year = int(data.get('year', 0))
        if not (2020 <= year <= 2035):
            errors.append('연도는 2020~2035 범위여야 합니다.')
    except (ValueError, TypeError):
        errors.append('올바른 연도를 입력해주세요.')

    try:
        month = int(data.get('month', 0))
        if not (1 <= month <= 12):
            errors.append('월은 1~12 범위여야 합니다.')
    except (ValueError, TypeError):
        errors.append('올바른 월을 입력해주세요.')

    employees = data.get('employees', [])
    if not isinstance(employees, list):
        errors.append('직원 목록이 올바르지 않습니다.')
    elif len(employees) < 2:
        errors.append('직원은 최소 2명 이상이어야 합니다.')
    elif len(employees) > 50:
        errors.append('직원은 최대 50명까지 지원합니다.')
    else:
        seen = set()
        for emp in employees:
            if not isinstance(emp, str) or len(emp.strip()) == 0:
                errors.append('직원 이름은 비어 있을 수 없습니다.')
                break
            if emp in seen:
                errors.append(f'중복된 직원 이름: {emp}')
                break
            seen.add(emp)

    cons = data.get('constraints', {})
    if isinstance(cons, dict):
        n = len(employees) if isinstance(employees, list) else 0
        checks = [
            ('doCount', 1, 28, 'D/O 횟수'),
            ('maxConsecutive', 1, 28, '최대 연속 근무'),
            ('maxConsecutiveOff', 1, 28, '최대 연속 휴무'),
            ('minWeekday', 1, n or 50, '평일 최소 근무'),
            ('minWeekend', 1, n or 50, '주말 최소 근무'),
            ('minWeekdayOff', 0, n or 50, '평일 최소 휴무'),
        ]
        for key, lo, hi, label in checks:
            try:
                v = int(cons.get(key, lo))
                if not (lo <= v <= hi):
                    errors.append(f'{label}은(는) {lo}~{hi} 범위여야 합니다.')
            except (ValueError, TypeError):
                errors.append(f'{label} 값이 올바르지 않습니다.')

    return errors


# ============================================================
# 파일 생성
# ============================================================

def _save_files(schedule, employees, year, month, cal_info, employee_do_counts, constraints, file_id):
    day_name_kr = {0: '월', 1: '화', 2: '수', 3: '목', 4: '금', 5: '토', 6: '일'}
    num_days = cal_info['numDays']
    days = list(range(1, num_days + 1))
    day_of_week = {int(k): v for k, v in cal_info['dayOfWeek'].items()}
    col_labels = [f"{d}({day_name_kr[day_of_week[d]]})" for d in days]

    df = pd.DataFrame(index=employees, columns=col_labels)
    for emp in employees:
        for i, d in enumerate(days):
            df.loc[emp, col_labels[i]] = schedule[emp].get(str(d), '')

    for emp in employees:
        vals = [schedule[emp].get(str(d), '') for d in days]
        df.loc[emp, '근무(W)'] = vals.count('W')
        df.loc[emp, '휴무(D/O)'] = vals.count('D/O')
        df.loc[emp, '연차(M/O)'] = vals.count('M/O')

    row = []
    for d in days:
        row.append(sum(1 for e in employees if schedule[e].get(str(d)) == 'W'))
    row += ['', '', '']
    df.loc['[근무인원]'] = row

    # CSV
    csv_name = f"schedule_{year}{month:02d}_{file_id}.csv"
    csv_path = os.path.join(OUTPUT_DIR, csv_name)
    df.to_csv(csv_path, encoding='utf-8-sig')
    csv_id = f"{file_id}_csv"
    file_store[csv_id] = (csv_path, time.time())

    # Excel
    excel_name = f"schedule_{year}{month:02d}_{file_id}.xlsx"
    excel_path = os.path.join(OUTPUT_DIR, excel_name)
    try:
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=f'{year}년{month}월')
            _style_excel(writer, f'{year}년{month}월')
        excel_id = f"{file_id}_excel"
        file_store[excel_id] = (excel_path, time.time())
    except Exception:
        excel_id = f"{file_id}_excel"
        file_store[excel_id] = (csv_path, time.time())

    return excel_id, csv_id


def _style_excel(writer, sheet_name):
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    ws = writer.sheets[sheet_name]

    fill_do = PatternFill(start_color='B4D7FF', end_color='B4D7FF', fill_type='solid')
    fill_mo = PatternFill(start_color='FFD6D6', end_color='FFD6D6', fill_type='solid')
    fill_header = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    font_header = Font(color='FFFFFF', bold=True, size=10)
    font_cell = Font(size=10)
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    center = Alignment(horizontal='center', vertical='center')

    for cell in ws[1]:
        cell.fill = fill_header
        cell.font = font_header
        cell.alignment = center
        cell.border = border

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.font = font_cell
            cell.alignment = center
            cell.border = border
            if cell.value == 'D/O':
                cell.fill = fill_do
            elif cell.value == 'M/O':
                cell.fill = fill_mo

    ws.column_dimensions['A'].width = 10
    for col_idx in range(2, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 8


# ============================================================
# 실행
# ============================================================

# DB 초기화
init_db()

if __name__ == '__main__':
    import sys
    is_dev = '--dev' in sys.argv
    port = int(os.environ.get('PORT', 5000))
    print("=" * 50)
    print("  직원 스케줄 생성기 (Commercial Edition)")
    print(f"  http://localhost:{port}")
    if is_dev:
        print("  [개발 모드]")
    print("=" * 50)
    app.run(host='0.0.0.0', port=port, debug=is_dev)
